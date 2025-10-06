import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from datetime import datetime, timedelta
import openpyxl
import os
import webbrowser

# Configuration for time ranges (can be easily modified)
TIME_RANGES = {
    'weekday': {'start': '12:00', 'end': '21:00'},  # Monday to Friday
    'weekend': {'start': '10:30', 'end': '21:00'}   # Saturday and Sunday
}

def calculate_hours_in_range(time_range, start_time, end_time):
    try:
        shift_start, shift_end = time_range.split("-")
        shift_start = datetime.strptime(shift_start.strip(), "%H:%M")
        
        # Handle cases where end time might be just a single digit (like "0" instead of "00:00")
        shift_end_str = shift_end.strip()
        if ":" not in shift_end_str:
            if len(shift_end_str) == 1:
                shift_end_str = f"0{shift_end_str}:00"
            else:
                shift_end_str = f"{shift_end_str}:00"
        
        shift_end = datetime.strptime(shift_end_str, "%H:%M")

        # Handle cases where shift end time is past midnight
        if shift_end <= shift_start:
            shift_end += timedelta(days=1)

        range_start = datetime.strptime(start_time, "%H:%M")
        range_end = datetime.strptime(end_time, "%H:%M")

        # Calculate overlap between shift and tip calculation range
        actual_start = max(shift_start, range_start)
        actual_end = min(shift_end, range_end)

        if actual_start < actual_end:
            hours = (actual_end - actual_start).total_seconds() / 3600  # Convert to hours
            print(f"Shift: {time_range}, Range: {start_time}-{end_time}, Overlap: {hours:.2f} hours")
            return hours
        return 0
    except (ValueError, AttributeError) as e:
        print(f"Error parsing time range '{time_range}': {e}")
        return 0

def process_file(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    end_of_roster_location = None
    names_and_hours = []

    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == "Break Notice":
                end_of_roster_location = (cell.row - 2, cell.column)  # Save the location (row, column)
                break
        if end_of_roster_location:
            break
    
    if end_of_roster_location:
        print(f"End of roster location: {end_of_roster_location}")
        # Extract names and hours
        start_row = 6
        end_row = end_of_roster_location[0]
        for row in range(start_row, end_row + 1):
            name = sheet.cell(row=row, column=1).value  # Names in column A
            hours = []
            for col in range(2, 9):  # Hours from column B to H (Monday to Sunday)
                cell_value = sheet.cell(row=row, column=col).value
                if cell_value:
                    cell_value_str = str(cell_value)
                    if "-" in cell_value_str:  # Check if it's a range
                        hours.append(cell_value_str[:9])  # Take the first 9 characters
                    else:
                        hours.append(cell_value_str)  # Otherwise, take the full value
                else:
                    hours.append("00:00-00:00")  # Default empty range
            names_and_hours.append((name, hours))

        # Calculate hours worked within specified ranges
        results = []
        daily_totals = [0] * 7  # Initialize daily totals for Monday to Sunday
        for name, hours in names_and_hours:
            daily_hours = []
            for i, time_range in enumerate(hours):
                if "-" in time_range and time_range != "00:00-00:00":
                    if i < 5:  # Monday to Friday (12:00-21:00)
                        worked_hours = calculate_hours_in_range(time_range, "12:00", "21:00")
                    elif i == 5 or i == 6:  # Saturday and Sunday (12:00-21:00) *
                        worked_hours = calculate_hours_in_range(time_range, "12:00", "21:00")
                    else:
                        worked_hours = 0
                else:
                    worked_hours = 0
                daily_hours.append(worked_hours)
                if i < len(daily_totals):  # Make sure we don't exceed the array bounds
                    daily_totals[i] += worked_hours  # Add to daily totals
            total_hours = sum(daily_hours)
            results.append((name, daily_hours, total_hours))

        # Print results
        for name, daily_hours, total_hours in results:
            print(f"Name: {name}, Daily Hours: {daily_hours}, Total Hours: {total_hours}")

        print("Daily Totals:")
        day_names = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
        for i, total in enumerate(daily_totals):
            print(f"{day_names[i]}: {total:.2f} hours")

        return results, daily_totals
    else:
        print("'Break Notice' not found in the file.")
        return [], [0] * 7

def display_results(results, daily_totals):
    result_window = tk.Tk()
    result_window.title("Tips Calculator - Select Days and Enter Tips")
    result_window.geometry("1000x700")
      # Day names
    day_names = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
    
    # Variables for tips
    tip_entries = []
    
    # Top frame for tips input
    top_frame = tk.Frame(result_window)
    top_frame.pack(fill=tk.X, padx=10, pady=10)
    
    # Tips input frame
    tips_frame = tk.LabelFrame(top_frame, text="Enter Tips for Each Day", font=("Arial", 12, "bold"))
    tips_frame.pack(fill=tk.X, pady=(0, 10))
    
    # Create a grid layout for better organization
    for i, day in enumerate(day_names):
        row = i // 4  # 4 days per row
        col = (i % 4) * 2
        
        tk.Label(tips_frame, text=f"{day}:", font=("Arial", 10)).grid(row=row, column=col, padx=5, pady=5, sticky="e")
        entry = tk.Entry(tips_frame, width=8, font=("Arial", 10))
        entry.grid(row=row, column=col+1, padx=5, pady=5)
        entry.insert(0, "0.00")  # Default value
        tip_entries.append(entry)    # Add instruction label
    instruction_text = "Enter tip amounts for each day (Kitchen gets 30%, remaining 70% split among staff by hours worked)"
    tk.Label(tips_frame, text=instruction_text, font=("Arial", 9), fg="gray").grid(row=2, column=0, columnspan=8, pady=5)
    
    # Calculate button
    def calculate_tips():
        # Determine selected days based on non-zero tip entries
        selected_days = []
        tips_per_day = {}
        
        for i in range(7):
            try:
                tip_amount = float(tip_entries[i].get()) if tip_entries[i].get() else 0
                if tip_amount < 0:
                    messagebox.showwarning("Invalid Input", f"Please enter a non-negative amount for {day_names[i]} tips.")
                    return
                if tip_amount > 0:  # Only include days with tips > 0
                    selected_days.append(i)
                    tips_per_day[i] = tip_amount
            except ValueError:
                messagebox.showerror("Invalid Input", f"Please enter a valid number for {day_names[i]} tips.")
                return
        
        if not selected_days:
            messagebox.showwarning("No Tips Entered", "Please enter tip amounts for at least one day.")
            return
        
        # Check if any selected day has zero hours
        zero_hour_days = []
        for day_idx in selected_days:
            if daily_totals[day_idx] == 0:
                zero_hour_days.append(day_names[day_idx])
        
        if zero_hour_days:
            messagebox.showwarning("No Hours Worked", f"No hours worked on: {', '.join(zero_hour_days)}. Please select different days or check your data.")
            return
        
        # Calculate rates and display results
        show_tip_calculation_results(results, daily_totals, selected_days, tips_per_day, day_names)
    
    tk.Button(top_frame, text="Calculate Tips", command=calculate_tips, font=("Arial", 12, "bold"), 
              bg="#4CAF50", fg="white", padx=20, pady=5).pack(pady=10)
    
    # Scrollable frame for hours table
    hours_frame = tk.LabelFrame(result_window, text="Hours Worked", font=("Arial", 12, "bold"))
    hours_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0, 10))
    
    # Create canvas and scrollbar for scrollable content
    canvas = tk.Canvas(hours_frame)
    scrollbar = ttk.Scrollbar(hours_frame, orient="vertical", command=canvas.yview)
    scrollable_frame = ttk.Frame(canvas)
    
    scrollable_frame.bind(
        "<Configure>",
        lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
    )
    
    canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
    canvas.configure(yscrollcommand=scrollbar.set)
    
    # Headers
    tk.Label(scrollable_frame, text="Name", font=("Arial", 12, "bold")).grid(row=0, column=0, padx=5, pady=5, sticky="w")
    for i, day in enumerate(day_names):
        tk.Label(scrollable_frame, text=day, font=("Arial", 12, "bold")).grid(row=0, column=i+1, padx=5, pady=5)
    tk.Label(scrollable_frame, text="Total", font=("Arial", 12, "bold")).grid(row=0, column=8, padx=5, pady=5)

    # Employee data
    for i, (name, daily_hours, total_hours) in enumerate(results, start=1):
        tk.Label(scrollable_frame, text=name, font=("Arial", 10)).grid(row=i, column=0, padx=5, pady=2, sticky="w")
        for j, hours in enumerate(daily_hours):
            tk.Label(scrollable_frame, text=f"{hours:.2f}", font=("Arial", 10)).grid(row=i, column=j + 1, padx=5, pady=2)
        tk.Label(scrollable_frame, text=f"{total_hours:.2f}", font=("Arial", 10)).grid(row=i, column=8, padx=5, pady=2)

    # Add daily totals at the bottom
    tk.Label(scrollable_frame, text="TOTAL", font=("Arial", 12, "bold")).grid(row=len(results) + 1, column=0, padx=5, pady=5, sticky="w")
    for j, total in enumerate(daily_totals):
        tk.Label(scrollable_frame, text=f"{total:.2f}", font=("Arial", 12, "bold")).grid(row=len(results) + 1, column=j + 1, padx=5, pady=5)
    
    # Grand total
    grand_total = sum(daily_totals)
    tk.Label(scrollable_frame, text=f"{grand_total:.2f}", font=("Arial", 12, "bold")).grid(row=len(results) + 1, column=8, padx=5, pady=5)

    canvas.pack(side="left", fill="both", expand=True)
    scrollbar.pack(side="right", fill="y")
    
    # Bind mousewheel to canvas
    def _on_mousewheel(event):
        canvas.yview_scroll(int(-1*(event.delta/120)), "units")
    canvas.bind_all("<MouseWheel>", _on_mousewheel)
    
    result_window.mainloop()

def show_tip_calculation_results(results, daily_totals, selected_days, tips_per_day, day_names):
    # Create new window for tip calculation results
    tip_window = tk.Toplevel()
    tip_window.title("Tips Calculator - Results")
    tip_window.geometry("1400x800")
    
    # Calculate kitchen share (30%) and staff share (70%) for each day
    kitchen_share = {}
    staff_share = {}
    hourly_rates = {}
    
    for day_idx in selected_days:
        total_tips_day = tips_per_day[day_idx]
        kitchen_share[day_idx] = total_tips_day * 0.30
        staff_share[day_idx] = total_tips_day * 0.70
        
        total_hours_day = daily_totals[day_idx]
        if total_hours_day > 0:
            # Calculate hourly rate based on the 70% staff share
            hourly_rates[day_idx] = staff_share[day_idx] / total_hours_day
        else:
            hourly_rates[day_idx] = 0

    # Grand totals
    total_tips_all = sum(tips_per_day.values())
    total_kitchen_all = sum(kitchen_share.values())
    total_staff_all = sum(staff_share.values())
    total_hours = sum(daily_totals[i] for i in selected_days)
    overall_rate = total_staff_all / total_hours if total_hours > 0 else 0
    
    # Top summary frame
    summary_frame = tk.LabelFrame(tip_window, text="Daily Summary", font=("Arial", 14, "bold"))
    summary_frame.pack(fill=tk.X, padx=10, pady=10)
    
    # Create a grid for daily summaries
    summary_row = 0
    for day_idx in selected_days:
        day_name = day_names[day_idx]
        total_tips = tips_per_day[day_idx]
        kitchen_tips = kitchen_share[day_idx]
        staff_tips = staff_share[day_idx]
        hours = daily_totals[day_idx]
        rate = hourly_rates[day_idx];
        
        tk.Label(summary_frame, text=f"{day_name}:", font=("Arial", 11, "bold")).grid(row=summary_row, column=0, padx=10, pady=2, sticky="w")
        tk.Label(summary_frame, text=f"Total: ${total_tips:.2f}", font=("Arial", 11)).grid(row=summary_row, column=1, padx=10, pady=2, sticky="w")
        tk.Label(summary_frame, text=f"Kitchen (30%): ${kitchen_tips:.2f}", font=("Arial", 11), fg="blue").grid(row=summary_row, column=2, padx=10, pady=2, sticky="w")
        tk.Label(summary_frame, text=f"Staff (70%): ${staff_tips:.2f}", font=("Arial", 11), fg="green").grid(row=summary_row, column=3, padx=10, pady=2, sticky="w")
        tk.Label(summary_frame, text=f"Hours: {hours:.2f}", font=("Arial", 11)).grid(row=summary_row, column=4, padx=10, pady=2, sticky="w")
        tk.Label(summary_frame, text=f"Rate: ${rate:.2f}/hr", font=("Arial", 11)).grid(row=summary_row, column=5, padx=10, pady=2, sticky="w")
        summary_row += 1    
    # Grand totals
    
    tk.Label(summary_frame, text="TOTAL:", font=("Arial", 12, "bold")).grid(row=summary_row, column=0, padx=10, pady=5, sticky="w")
    tk.Label(summary_frame, text=f"${total_tips_all:.2f}", font=("Arial", 12, "bold")).grid(row=summary_row, column=1, padx=10, pady=5, sticky="w")
    tk.Label(summary_frame, text=f"${total_kitchen_all:.2f}", font=("Arial", 12, "bold"), fg="blue").grid(row=summary_row, column=2, padx=10, pady=5, sticky="w")
    tk.Label(summary_frame, text=f"${total_staff_all:.2f}", font=("Arial", 12, "bold"), fg="green").grid(row=summary_row, column=3, padx=10, pady=5, sticky="w")
    tk.Label(summary_frame, text=f"{total_hours:.2f}", font=("Arial", 12, "bold")).grid(row=summary_row, column=4, padx=10, pady=5, sticky="w")
    tk.Label(summary_frame, text=f"${overall_rate:.2f}/hr", font=("Arial", 12, "bold")).grid(row=summary_row, column=5, padx=10, pady=5, sticky="w")
      # Print button frame
    button_frame = tk.Frame(tip_window)
    button_frame.pack(fill=tk.X, padx=10, pady=5)
    
    def generate_html_report(results, daily_totals, selected_days, tips_per_day, day_names, kitchen_share, staff_share, hourly_rates, total_tips_all, total_kitchen_all, total_staff_all, total_hours, overall_rate):
        try:
            # Get today's date for filename
            today = datetime.now().strftime("%Y-%m-%d")
              # Ask user where to save the HTML report
            html_path = filedialog.asksaveasfilename(
                title="Save Tips Calculation Report",
                defaultextension=".html",
                filetypes=[("HTML files", "*.html"), ("All files", "*.*")],
                initialfile="results.html"
            )
            
            if not html_path:
                return  # User cancelled
            
            # Generate HTML content using the working test logic
            date_str = datetime.now().strftime("%B %d, %Y at %I:%M %p")
            
            html_content = f"""<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <title>Tips Calculation Results</title>
    <style>
        body {{
            font-family: Arial, sans-serif;
            margin: 20px;
            color: #333;
        }}
        .header {{
            text-align: center;
            margin-bottom: 30px;
        }}
        .header h1 {{
            color: #2c3e50;
            margin-bottom: 10px;
        }}
        .date {{
            color: #7f8c8d;
            font-size: 14px;
        }}
        .summary {{
            margin-bottom: 30px;
        }}
        .summary h2 {{
            color: #34495e;
            border-bottom: 2px solid #3498db;
            padding-bottom: 5px;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            margin-bottom: 20px;
        }}
        th, td {{
            border: 1px solid #bdc3c7;
            padding: 8px;
            text-align: center;
        }}
        th {{
            background-color: #34495e;
            color: white;
            font-weight: bold;
        }}
        .total-row {{
            background-color: #ecf0f1;
            font-weight: bold;
        }}
        .kitchen {{
            color: #2980b9;
        }}
        .staff {{
            color: #27ae60;
        }}
        .employee-total {{
            color: #27ae60;
            font-weight: bold;
        }}
        @media print {{
            body {{
                margin: 10px;
            }}
            .no-print {{
                display: none;
            }}
        }}
        .print-note {{
            background-color: #f8f9fa;
            border: 1px solid #dee2e6;
            border-radius: 5px;
            padding: 15px;
            margin: 20px 0;
            text-align: center;
        }}
        .print-button {{
            background-color: #007bff;
            color: white;
            padding: 10px 20px;
            border: none;
            border-radius: 5px;
            cursor: pointer;
            font-size: 16px;
            margin: 10px;
        }}
        .print-button:hover {{
            background-color: #0056b3;
        }}
    </style>
</head>
<body>
    <div class="header">
        <h1>TIPS CALCULATION RESULTS</h1>
        <div class="date">Generated on: {date_str}</div>
    </div>
    
    <div class="print-note no-print">
        <p><strong>To save as PDF:</strong> Use your browser's Print function and select "Save as PDF" as the destination.</p>
        <button class="print-button" onclick="window.print()">Print / Save as PDF</button>
    </div>

    <div class="summary">
        <h2>DAILY SUMMARY</h2>
        <table>
            <thead>
                <tr>
                    <th>Day</th>
                    <th>Total Tips</th>
                    <th>Kitchen (30%)</th>
                    <th>Staff (70%)</th>
                    <th>Hours</th>
                    <th>Rate/Hour</th>
                </tr>
            </thead>
            <tbody>"""
            
            # Add daily summary rows - using the variables from the outer scope
            for day_idx in selected_days:
                day_name = day_names[day_idx]
                total_tips = tips_per_day[day_idx]
                kitchen_tips = kitchen_share[day_idx]
                staff_tips = staff_share[day_idx]
                hours = daily_totals[day_idx]
                rate = hourly_rates[day_idx]
                
                html_content += f"""
                <tr>
                    <td>{day_name}</td>
                    <td>${total_tips:.2f}</td>
                    <td class="kitchen">${kitchen_tips:.2f}</td>
                    <td class="staff">${staff_tips:.2f}</td>
                    <td>{hours:.2f}</td>
                    <td>${rate:.2f}</td>
                </tr>"""
            
            # Add totals row
            html_content += f"""
                <tr class="total-row">
                    <td><strong>TOTAL</strong></td>
                    <td><strong>${total_tips_all:.2f}</strong></td>
                    <td class="kitchen"><strong>${total_kitchen_all:.2f}</strong></td>
                    <td class="staff"><strong>${total_staff_all:.2f}</strong></td>
                    <td><strong>{total_hours:.2f}</strong></td>
                    <td><strong>${overall_rate:.2f}</strong></td>
                </tr>
            </tbody>
        </table>
    </div>

    <div class="summary">
        <h2>INDIVIDUAL EMPLOYEE RESULTS</h2>
        <table>
            <thead>
                <tr>
                    <th>Employee</th>"""
            
            # Add day headers
            for day_idx in selected_days:
                html_content += f"<th>{day_names[day_idx]}</th>"
            html_content += "<th>Total Tips</th></tr></thead><tbody>"
            
            # Add employee data
            for name, daily_hours, total_hours in results:
                employee_total_tips = 0
                html_content += f"<tr><td>{name if name else 'Unknown'}</td>"
                
                for day_idx in selected_days:
                    hours = daily_hours[day_idx]
                    rate = hourly_rates[day_idx]
                    tips = hours * rate
                    employee_total_tips += tips
                    
                    if hours > 0:
                        html_content += f"<td>{hours:.1f}h<br>${tips:.2f}</td>"
                    else:
                        html_content += "<td>-</td>"
                
                html_content += f'<td class="employee-total">${employee_total_tips:.2f}</td></tr>'
            
            # Close the HTML
            html_content += """
            </tbody>
        </table>
    </div>
    
    <script>
        // Optional: Auto-focus print dialog (some browsers may block this)
        // window.onload = function() { window.print(); }
    </script>
</body>
</html>"""
            
            # Write HTML file
            with open(html_path, 'w', encoding='utf-8') as f:
                f.write(html_content)
            
            # Show success message and ask if user wants to open the HTML
            response = messagebox.askyesno(
                "Report Generated", 
                f"Tips calculation report has been saved to:\n{html_path}\n\n"
                f"You can:\n"
                f"• Open it in your browser and use Print → Save as PDF\n"
                f"• Email the HTML file directly\n"
                f"• Print it directly from your browser\n\n"
                f"Would you like to open it now?"
            )
            
            if response:
                try:
                    # Open in default browser
                    webbrowser.open(f'file://{os.path.abspath(html_path)}')
                except Exception as e:
                    messagebox.showinfo(
                        "Report Saved", 
                        f"Report saved successfully to:\n{html_path}\n\n"
                        f"Please open it manually in your web browser."
                    )                
        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate report: {str(e)}")
    
    tk.Button(button_frame, text="Generate HTML Report", command=lambda: generate_html_report(results, daily_totals, selected_days, tips_per_day, day_names, kitchen_share, staff_share, hourly_rates, total_tips_all, total_kitchen_all, total_staff_all, total_hours, overall_rate), 
              font=("Arial", 12, "bold"), bg="#FF9800", fg="white", padx=20, pady=5).pack(side=tk.LEFT, padx=5)
    
    def export_to_text_file(results, daily_totals, selected_days, tips_per_day, day_names, kitchen_share, staff_share, hourly_rates, total_tips_all, total_kitchen_all, total_staff_all, total_hours, overall_rate):
        try:
            # Get today's date for filename
            today = datetime.now().strftime("%Y-%m-%d")
            # Ask user where to save the text file
            file_path = filedialog.asksaveasfilename(
                title="Save Report as Text File",
                defaultextension=".txt",
                filetypes=[("Text files", "*.txt"), ("All files", "*.*")],
                initialfile="results.txt"
            )
            
            if not file_path:
                return  # User cancelled

            # Generate file content
            content = []
            date_str = datetime.now().strftime("%B %d, %Y at %I:%M %p")
            
            content.append("TIPS CALCULATION RESULTS")
            content.append(f"Generated on: {date_str}")
            content.append("="*40)
            content.append("\nDAILY SUMMARY")
            content.append("-"*40)

            # Daily Summary
            for day_idx in selected_days:
                day_name = day_names[day_idx]
                total_tips = tips_per_day[day_idx]
                kitchen_tips = kitchen_share[day_idx]
                staff_tips = staff_share[day_idx]
                hours = daily_totals[day_idx]
                rate = hourly_rates[day_idx]
                content.append(f"{day_name}:")
                content.append(f"  Total Tips: ${total_tips:9.2f}")
                content.append(f"  Kitchen (30%): ${kitchen_tips:9.2f}")
                content.append(f"  Staff (70%):   ${staff_tips:9.2f}")
                content.append(f"  Total Hours: {hours:6.2f}")
                content.append(f"  Rate/Hour:   ${rate:9.2f}")
                content.append("")

            # Grand Totals
            content.append("TOTALS:")
            content.append(f"  Total Tips: ${total_tips_all:9.2f}")
            content.append(f"  Kitchen (30%): ${total_kitchen_all:9.2f}")
            content.append(f"  Staff (70%):   ${total_staff_all:9.2f}")
            content.append(f"  Total Hours: {total_hours:6.2f}")
            content.append(f"  Overall Rate:  ${overall_rate:9.2f}")

            content.append("\n\n" + "="*40)
            content.append("INDIVIDUAL EMPLOYEE RESULTS")
            content.append("-"*40)

            # Employee data
            header = f"{'Employee':<20}"
            for day_idx in selected_days:
                header += f" | {day_names[day_idx]:^15}"
            header += f" | {'Total Tips':>12}"
            content.append(header)
            content.append("-" * len(header))

            for name, daily_hours, _ in results:
                employee_total_tips = 0
                row_str = f"{name if name else 'Unknown':<20}"
                
                tips_by_day = {}
                for day_idx in selected_days:
                    hours = daily_hours[day_idx]
                    rate = hourly_rates[day_idx]
                    tips = hours * rate
                    employee_total_tips += tips
                    tips_by_day[day_idx] = (hours, tips)

                for day_idx in selected_days:
                    hours, tips = tips_by_day[day_idx]
                    if hours > 0:
                        row_str += f" | {f'{hours:.1f}h / ${tips:.2f}':^15}"
                    else:
                        row_str += f" | {'-':^15}"
                
                row_str += f" | ${employee_total_tips:>11.2f}"
                content.append(row_str)

            # Write to file
            with open(file_path, 'w', encoding='utf-8') as f:
                f.write('\n'.join(content))
            
            messagebox.showinfo(
                "Report Saved", 
                f"Printable report has been saved to:\n{file_path}"
            )
                
        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate report: {str(e)}")
    
    tk.Button(button_frame, text="Save as Text File", command=lambda: export_to_text_file(results, daily_totals, selected_days, tips_per_day, day_names, kitchen_share, staff_share, hourly_rates, total_tips_all, total_kitchen_all, total_staff_all, total_hours, overall_rate), 
              font=("Arial", 11), bg="#28a745", fg="white", padx=15, pady=5).pack(side=tk.LEFT, padx=5)
    
    tk.Button(button_frame, text="Close", command=tip_window.destroy, 
              font=("Arial", 12), bg="#757575", fg="white", padx=20, pady=5).pack(side=tk.RIGHT, padx=5)
    
    # Scrollable frame for detailed results
    results_frame = tk.LabelFrame(tip_window, text="Individual Employee Results", font=("Arial", 14, "bold"))
    results_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0, 10))
    
    # Create canvas and scrollbar
    canvas = tk.Canvas(results_frame)
    scrollbar = ttk.Scrollbar(results_frame, orient="vertical", command=canvas.yview)
    scrollable_frame = ttk.Frame(canvas)
    
    scrollable_frame.bind(
        "<Configure>",
        lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
    )
    
    canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
    canvas.configure(yscrollcommand=scrollbar.set)
      # Headers
    col = 0
    tk.Label(scrollable_frame, text="Name", font=("Arial", 12, "bold")).grid(row=0, column=col, padx=5, pady=5, sticky="w")
    col += 1
    
    for day_idx in selected_days:
        tk.Label(scrollable_frame, text=f"{day_names[day_idx]}", font=("Arial", 11, "bold")).grid(row=0, column=col, padx=3, pady=5)
        col += 1
        tk.Label(scrollable_frame, text="Hours", font=("Arial", 10, "bold")).grid(row=1, column=col-1, padx=3, pady=2)
        tk.Label(scrollable_frame, text="Tips", font=("Arial", 10, "bold")).grid(row=2, column=col-1, padx=3, pady=2)
    
    tk.Label(scrollable_frame, text="Total Tips", font=("Arial", 12, "bold")).grid(row=0, column=col, padx=5, pady=5)
      # Employee data
    for row_idx, (name, daily_hours, total_hours) in enumerate(results, start=3):
        col = 0
        tk.Label(scrollable_frame, text=name, font=("Arial", 10)).grid(row=row_idx, column=col, padx=5, pady=3, sticky="w")
        col += 1
        
        employee_total_tips = 0
        
        for day_idx in selected_days:
            hours = daily_hours[day_idx]
            rate = hourly_rates[day_idx]  # This is already calculated from 70% staff share
            tips = hours * rate
            employee_total_tips += tips
            
            # Create a mini-frame for each day's data
            day_frame = tk.Frame(scrollable_frame)
            day_frame.grid(row=row_idx, column=col, padx=3, pady=1)
            
            tk.Label(day_frame, text=f"{hours:.1f}h", font=("Arial", 9)).pack()
            tk.Label(day_frame, text=f"${tips:.2f}", font=("Arial", 9, "bold"), fg="green").pack()
            
            col += 1
        
        tk.Label(scrollable_frame, text=f"${employee_total_tips:.2f}", font=("Arial", 11, "bold"), fg="darkgreen").grid(row=row_idx, column=col, padx=5, pady=3)
    
    canvas.pack(side="left", fill="both", expand=True)
    scrollbar.pack(side="right", fill="y")
    
    # Bind mousewheel to canvas
    def _on_mousewheel(event):
        canvas.yview_scroll(int(-1*(event.delta/120)), "units")
    canvas.bind_all("<MouseWheel>", _on_mousewheel)

def select_file():
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if file_path:
        file_name = file_path.split('/')[-1]  # Extract the file name
        print(f"Selected file: {file_name}")
        results, daily_totals = process_file(file_path)
        if results:  # Only display if we have results
            display_results(results, daily_totals)
        root.destroy()  # Close the GUI

# Create the main window
root = tk.Tk()
root.title("Tips Calculator - Select Excel File")
root.geometry("600x500")

# Main frame with scrollbar for instructions
main_frame = tk.Frame(root)
main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

# Title
title_label = tk.Label(main_frame, text="Tips Calculator", font=("Arial", 16, "bold"))
title_label.pack(pady=10)

# Instructions frame
instructions_frame = tk.LabelFrame(main_frame, text="How to Export Roster from Alkimii", font=("Arial", 12, "bold"))
instructions_frame.pack(fill=tk.X, pady=10)

# Create canvas and scrollbar for instructions
canvas = tk.Canvas(instructions_frame, height=200)
scrollbar = ttk.Scrollbar(instructions_frame, orient="vertical", command=canvas.yview)
scrollable_frame = ttk.Frame(canvas)

scrollable_frame.bind(
    "<Configure>",
    lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
)

canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
canvas.configure(yscrollcommand=scrollbar.set)

# Add step-by-step instructions
instructions = [
    "1. Go to Alkimii and navigate to your roster",
    "2. Select the week you want to calculate tips for",
    "3. Select 'Export' from the top right menu (Small Arrow and Circle Icon)",
    "4. Download the exported file",
    "5. Open the downloaded file in Excel",
    "6. Remove any Shifts that are Reception/Kitchen (if any)",
    "7. Use that Excel file with this Tips Calculator",
    "8. Ensure all names are on the Calculator",
    "8. Proof read any results with the Roster and Calculator, if there is any issues tips will have to be done manually"
]

for i, instruction in enumerate(instructions):
    tk.Label(scrollable_frame, text=instruction, font=("Arial", 10), anchor="w", justify="left").pack(fill=tk.X, padx=10, pady=2)


canvas.pack(side="left", fill="both", expand=True)
scrollbar.pack(side="right", fill="y")

# File selection section
file_frame = tk.LabelFrame(main_frame, text="Select Roster File", font=("Arial", 12, "bold"))
file_frame.pack(fill=tk.X, pady=10)

instruction_label = tk.Label(file_frame, text="Select the exported Excel file from Alkimii", font=("Arial", 12))
instruction_label.pack(pady=5)

# Create a button to open the file dialog
select_button = tk.Button(file_frame, text="Select Excel File", command=select_file, 
                         font=("Arial", 12), bg="#2196F3", fg="white", padx=30, pady=10)
select_button.pack(pady=20)

# Add exit button
exit_button = tk.Button(main_frame, text="Exit", command=root.destroy, 
                       font=("Arial", 10), padx=20, pady=5)
exit_button.pack(pady=10)

# Run the GUI event loop
root.mainloop()
